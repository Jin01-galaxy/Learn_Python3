from slackbot.bot import respond_to
from datetime import datetime
import openpyxl

@respond_to('出勤')
def punch_in(message):
    print('出勤時刻を登録')
    timestamp = datetime.now()
    message.send('出勤時刻は、{}です。'.format(timestamp.strftime('%H:%M')))
    wb = openpyxl.load_workbook('勤怠管理.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    ws[max_row + 1][0].value = timestamp.strftime('%y/%m/%d')
    ws[max_row + 1][1].value = timestamp.strftime('%H:%M')
    wb.save('勤怠管理.xlsx')
    print('登録完了しました。')
    
@respond_to('退勤')
def punch_out(message):
    print('退勤時刻を登録')
    timestamp = datetime.now()
    message.send('退勤時刻は、{}です。'.format(timestamp.strftime('%H:%M')))
    wb = openpyxl.load_workbook('勤怠管理.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    ws[max_row][2].value = timestamp.strftime('%H:%M')
    wb.save('勤怠管理.xlsx')
    print('登録完了しました。')
    
    